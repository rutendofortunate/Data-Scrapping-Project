import os
import pandas as pd
import tabula
from openpyxl import Workbook

# Specify the directory containing the PDF files
pdf_dir = r'C:\Users\horon\OneDrive\Desktop\IH Securities\Price Sheet\PDFs'
excel_output = 'output.xlsx'

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Iterate over all PDF files in the specified directory
for pdf_file in os.listdir(pdf_dir):
    if pdf_file.endswith('.pdf'):
        # Full path to your PDF
        full_pdf_path = os.path.join(pdf_dir, pdf_file)
        
        # Use tabula to read tables from the PDF file
        # This might need adjustment depending on the PDF's structure
        df_list = tabula.read_pdf(full_pdf_path, pages='all', multiple_tables=True)
        
        for idx, df in enumerate(df_list):
            # Create a new sheet for each table in the PDF
            # If only one table per PDF is expected, you might adjust this part
            sheet_name = f"{pdf_file}_{idx}"
            ws = wb.create_sheet(title=sheet_name[:30])  # Sheet name limited to 31 characters
            
            # Convert DataFrame to rows and write to the sheet
            for r_idx, row in enumerate(df.to_numpy(), start=1):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook
wb.save(excel_output)
